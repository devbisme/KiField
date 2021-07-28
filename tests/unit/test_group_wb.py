import openpyxl as pyxl
from kifield import kifield


def test_groups():
    wb = pyxl.Workbook()
    ws = wb.active
    header = ("Ref", "x", "y", "z")
    ws.append(header)
    ws.append(("C1", "1", "1", "1"))
    ws.append(("C2", "1", "1", "1"))
    ws.append(("C3", "1", "1", "1"))

    wb = kifield.group_wb(wb)
    ws = wb.active

    assert ws.max_row == 2
    assert ws.max_column == 4

    values = tuple(ws.values)
    assert values[0] == header
    assert values[1] == ("C1-C3", "1", "1", "1")

def test_groups2():
    wb = pyxl.Workbook()
    ws = wb.active
    header = ("Ref", "x", "y", "z")
    ws.append(header)
    ws.append(("C1", "1", "1", "1"))
    ws.append(("R3", "2", "1", "1"))
    ws.append(("R5", "2", "1", "1"))
    ws.append(("X1", "3", "1", "1"))
    ws.append(("X2", "1", "3", "1"))
    ws.append(("X3", "1", "1", "3"))

    wb = kifield.group_wb(wb)
    ws = wb.active

    assert ws.max_row == 6
    assert ws.max_column == 4

    values = tuple(ws.values)
    assert values[0] == header
    assert values[1] == ("C1", "1", "1", "1")
    assert values[2] == ("R3, R5", "2", "1", "1")
    assert values[3] == ("X1", "3", "1", "1")
    assert values[4] == ("X2", "1", "3", "1")
    assert values[5] == ("X3", "1", "1", "3")


# Test for no-range flag
def test_groups3():
    wb = pyxl.Workbook()
    ws = wb.active
    header = ("Ref", "x", "y", "z")
    ws.append(header)
    ws.append(("C1", "1", "1", "1"))
    ws.append(("C2", "1", "1", "1"))
    ws.append(("C3", "1", "1", "1"))

    wb = kifield.group_wb(wb, no_range=True)
    ws = wb.active

    assert ws.max_row == 2
    assert ws.max_column == 4

    values = tuple(ws.values)
    assert values[0] == header
    assert values[1] == ("C1, C2, C3", "1", "1", "1")
