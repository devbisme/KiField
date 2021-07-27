# MIT License / Copyright (c) 2016-2021 by Dave Vandenbout.


from __future__ import absolute_import, division, print_function, unicode_literals

import csv
import logging
import operator
import os
import pdb
import random
import re
import string
import sys
from builtins import int, open, range, str

import openpyxl as pyxl
from future import standard_library

standard_library.install_aliases()


def csvfile_to_wb(csv_filename):
    """Open a CSV file and return an openpyxl workbook."""

    with open(csv_filename) as csv_file:
        reader = csv.reader(csv_file)
        wb = pyxl.Workbook()
        ws = wb.active
        for row_index, row in enumerate(reader, 1):
            for column_index, cell in enumerate(row, 1):
                if cell not in ("", None):
                    ws.cell(row=row_index, column=column_index).value = cell
    return wb


def wb_to_csvfile(wb, csv_filename):
    """Save an openpyxl workbook as a CSV file."""

    ws = wb.active
    mode = "w"
    if sys.version_info.major < 3:
        mode += "b"
    with open(csv_filename, mode) as csv_file:
        writer = csv.writer(csv_file, lineterminator="\n")
        for row in ws.rows:
            writer.writerow([cell.value for cell in row])


def random_string(length):
    return "".join(
        [random.choice(string.ascii_letters + string.digits) for n in range(1, length)]
    )


random.seed()
# Get worksheet.
wb = csvfile_to_wb(sys.argv[1])
ws = wb.active
# Find number of columns to add.
num_addtl_cols = random.randint(1, 4)
# Create sorted column headers.
col_hdrs = sorted(
    set(["zzz" + random_string(random.randint(1, 10)) for _ in range(num_addtl_cols)])
)
num_addtl_cols = len(col_hdrs)
# Add column headers.
start_column = ws.max_column + 1
for c, h in enumerate(col_hdrs):
    ws.cell(row=1, column=start_column + c).value = h
# Add random data to the first row so the columns don't disappear during extraction.
for c in range(start_column, ws.max_column + 1):
    ws.cell(row=2, column=c).value = random_string(random.randint(5, 20))
# Add random data to each new column.
for c in range(2, ws.max_column + 1):
    for r in range(2, ws.max_row + 1):
        if random.randint(0, 10) < 1:
            ws.cell(row=r, column=c).value = random_string(random.randint(5, 20))
# Write new spreadsheet.
wb_to_csvfile(wb, sys.argv[2])
