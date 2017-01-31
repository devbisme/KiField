# MIT license
# 
# Copyright (C) 2016 by XESS Corp.
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.


from __future__ import print_function
from __future__ import unicode_literals
from __future__ import division
from __future__ import absolute_import
from builtins import open
from builtins import range
from builtins import int
from builtins import str
from future import standard_library
standard_library.install_aliases()

import sys
import os
import re
import operator
import csv
import string
import random
import logging
import openpyxl as pyxl
import pdb



def csvfile_to_wb(csv_filename):
    '''Open a CSV file and return an openpyxl workbook.'''

    with open(csv_filename) as csv_file:
        reader = csv.reader(csv_file)
        wb = pyxl.Workbook()
        ws = wb.active
        for row_index, row in enumerate(reader, 1):
            for column_index, cell in enumerate(row, 1):
                if cell not in ('', None):
                    ws.cell(row=row_index, column=column_index).value = cell
    return wb


def reorder_ws_columns(ws, new_headers):
    headers = [ws.cell(row=1, column=c).value for c in range(1,ws.max_column+1)]
    new_order = [headers.index(new_h)+1 for new_h in new_headers]
    for r in range(1,ws.max_row+1):
        vals = [ws.cell(row=r,column=c).value for c in new_order]
        for c in range(1,ws.max_column+1):
            ws.cell(row=r, column=c).value = vals[c-1]


def wb_to_csvfile(wb, csv_filename):
    '''Save an openpyxl workbook as a CSV file.'''

    ws = wb.active
    mode = 'w'
    if sys.version_info.major < 3:
        mode += 'b'
    with open(csv_filename, mode) as csv_file:
        writer = csv.writer(csv_file, lineterminator='\n')
        for row in ws.rows:
            writer.writerow([cell.value for cell in row])

def random_string(length):
    return ''.join([random.choice(string.ascii_letters + string.digits) for n in range(1,length)])


random.seed()
# Get worksheet.
wb = csvfile_to_wb(sys.argv[1])
ws = wb.active
# Add column headers.
col_hdrs = ['description', 'keywords', 'docfile']
start_column = ws.max_column+1
for c,h in enumerate(col_hdrs):
    ws.cell(row=1, column=start_column+c).value = h
# Add random data to the first row so the columns don't disappear during extraction.
for c in range(start_column, ws.max_column+1):
    ws.cell(row=2, column=c).value = random_string(random.randint(5,20))
# Add random data to each new column.
for c in range(start_column, ws.max_column+1):
    for r in range(2, ws.max_row+1):
        if random.randint(0,10) < 10:
            ws.cell(row=r, column=c).value = random_string(random.randint(5,20))
# Reorder columns.
new_headers = sorted([ws.cell(row=1, column=c).value for c in range(2,ws.max_column+1)])
new_headers.insert(0,ws.cell(row=1,column=1).value)
reorder_ws_columns(ws, new_headers) 
# Write new spreadsheet.
wb_to_csvfile(wb, sys.argv[2])
