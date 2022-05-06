# -*- coding: utf-8 -*-

# MIT License / Copyright (c) 2021 by Dave Vandenbout.

from __future__ import absolute_import, division, print_function, unicode_literals

import csv
import operator
import os
import os.path
import re
from builtins import bytes, dict, int, map, open, range, str
from copy import deepcopy
from difflib import get_close_matches
from pprint import pprint

import openpyxl as pyxl
from future import standard_library

from .common import *
from .dcm import Component, Dcm
from .sch import sch_field_id_to_name, Schematic, Schematic_V6
from .schlib import SchLib, SchLib_V6

standard_library.install_aliases()


logger = logging.getLogger("kifield")

# Assign some names to the unnamed fields in a schematic or library component.
lib_field_id_to_name = {"0": "prefix", "1": "value", "2": "footprint", "3": "datasheet"}
lib_field_name_to_id = {v: k for k, v in lib_field_id_to_name.items()}
dcm_field_names = ["description", "keywords", "docfile"]

INVIS_PREFIX = "[I]"
VISIBLE_PREFIX = "[V]"


def csvfile_to_wb(csv_filename):
    """Open a CSV file and return an openpyxl workbook."""

    logger.log(
        DEBUG_DETAILED,
        "Converting CSV file {} into an XLSX workbook.".format(csv_filename),
    )

    with open(csv_filename) as csv_file:
        dialect = csv.Sniffer().sniff(csv_file.read())
        if USING_PYTHON2:
            for attr in dir(dialect):
                a = getattr(dialect, attr)
                if type(a) == unicode:
                    setattr(dialect, attr, a.encode("utf-8"))
        csv_file.seek(0)
        reader = csv.reader(csv_file, dialect)
        wb = pyxl.Workbook()
        ws = wb.active
        for row_index, row in enumerate(reader, 1):
            for column_index, cell in enumerate(row, 1):
                if cell not in ("", None):
                    ws.cell(row=row_index, column=column_index).value = cell
    return (wb, dialect)


def wb_to_csvfile(wb, csv_filename, dialect):
    """Save an openpyxl workbook as a CSV file."""

    logger.log(
        DEBUG_DETAILED,
        "Converting an XLSX workbook and saving as CSV file {}.".format(csv_filename),
    )
    ws = wb.active
    mode = "w"
    if USING_PYTHON2:
        mode += "b"
    with open(csv_filename, mode) as csv_file:
        writer = csv.writer(csv_file, dialect=dialect, lineterminator="\n")
        for row in ws.rows:
            writer.writerow([cell.value for cell in row])


def group_wb(wb, no_range=False):
    """Group lines that have the same column values in a openpyxl workbook.
    Headers are expected on the first row and references are expected in the
    first column."""

    ws = wb.active
    values = tuple(ws.values)

    try:
        header = values[0]
    except IndexError:
        # No header, so don't even try to ungroup the workbook.
        return wb

    unique_rows = []
    references = []
    for row in values[1:]:
        column_values = row[1:]
        reference = row[0]
        if column_values in unique_rows:
            index = unique_rows.index(column_values)
            references[index].append(reference)
        else:
            unique_rows.append(column_values)
            references.append([reference])

    grouped_rows = []
    for i, ref in enumerate(references):
        # If no_range flag, do the collapse to ensure sorting, then explode, then join to string
        if no_range:
            collapsed_refs = collapse(ref)
            exploded_refs = explode(collapsed_refs)
            joined_refs = (', ').join(exploded_refs)
            grouped_rows.append((joined_refs,) + unique_rows[i])
        else:
            grouped_rows.append((collapse(ref),) + unique_rows[i])

    grouped_wb = pyxl.Workbook()
    grouped_ws = grouped_wb.active
    grouped_ws.append(header)

    for row in grouped_rows:
        grouped_ws.append(row)

    return grouped_wb


def ungroup_wb(wb):
    """Ungroup lines that have collapsed references."""
    # return wb

    ws = wb.active
    values = tuple(ws.values)

    try:
        header = values[0]
    except IndexError:
        # No header, so don't even try to ungroup the workbook.
        return wb

    ungrouped_wb = pyxl.Workbook()
    ungrouped_ws = ungrouped_wb.active
    ungrouped_ws.append(header)

    for row in values[1:]:
        column_values = row[1:]
        reference = row[0]
        for ref in explode(reference):
            ungrouped_row = ((ref,) + column_values)
            ungrouped_ws.append(ungrouped_row)

    return ungrouped_wb


class FieldExtractionError(Exception):
    pass


class FindLabelError(Exception):
    pass


def find_header(ws):
    """Find the spreadsheet row that most likely contains the field headers."""

    # Look for the first occurrence of the row with the most entries.
    # That's probably the header row.
    max_width = 0
    header_row_num = 0
    header = []
    r = enumerate(ws.rows, 1)
    for row_num, row in r:
        width = len(row) - [c.value for c in row].count(None)
        if width > max_width:
            max_width = width
            header_row_num = row_num
            header = [cell for cell in row]

    logger.log(
        DEBUG_DETAILED,
        "Header on row {}: {}.".format(header_row_num, [c.value for c in header]),
    )
    return header_row_num, header


def lc_get_close_matches(lbl, possibilities, num_matches=3, cutoff=0.6):
    """Return list of closest matches to lbl from possibilities (case-insensitive)."""

    # Strip any non-strings so str.lower() doesn't crash.
    possibilities = [p for p in possibilities if isinstance(p, basestring)]

    if USING_PYTHON2:
        lc_lbl = str.lower(unicode(lbl))
        lc_possibilities = [str.lower(unicode(p)) for p in possibilities]
    else:
        lc_lbl = str.lower(lbl)
        lc_possibilities = [str.lower(p) for p in possibilities]
    lc_matches = get_close_matches(lc_lbl, lc_possibilities, num_matches, cutoff)
    return [possibilities[lc_possibilities.index(m)] for m in lc_matches]


def find_header_column(header, lbl):
    """Find the field header column containing the closest match to the given label."""

    header_labels = [cell.value for cell in header]
    lbl_match = lc_get_close_matches(lbl, header_labels, 1, 0.0)[0]
    for cell in header:
        if str(cell.value).lower() == lbl_match.lower():
            logger.log(
                DEBUG_OBSESSIVE,
                "Found {} on header column {}.".format(lbl, cell.column),
            )
            return cell.column, lbl_match
    raise FindLabelError("{} not found in spreadsheet".format(lbl))


def cull_list(fields, inc_fields=None, exc_fields=None):
    """Update the list by keeping only items in inc_fields and deleting items in exc_fields."""

    # Remove any fields not in include list unless this list is empty, in which case keep all fields.
    try:
        if len(inc_fields) > 0:
            for field in fields[:]:
                if len(lc_get_close_matches(field, inc_fields, 1, 0.6)) == 0:
                    fields.remove(field)
    except TypeError:
        # inc_fields was not a list, so do nothing to the field list.
        pass

    # Delete these fields unless list is empty, in which case delete none of them.
    try:
        if len(exc_fields) > 0:
            for field in fields[:]:
                if len(lc_get_close_matches(field, exc_fields, 1, 0.6)) != 0:
                    fields.remove(field)
    except TypeError:
        # exc_fields was not a list, so do nothing to the field list.
        pass


def extract_part_fields_from_wb(
    wb, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from an XLSX workbook."""

    wb = ungroup_wb(wb)

    part_fields = {}  # Start with an empty part dictionary.

    try:
        ws = wb.active  # Get the active worksheet from the workbook.

        # Find the header with the part field labels.
        header_row, header = find_header(ws)

        # Find the column with the part references.
        refs_c, refs_lbl = find_header_column(header, "refs")

        # Get the list of part references.
        refs = [r.value for r in list(ws.columns)[refs_c - 1][header_row:]]

        # Make a dict of spreadsheet column indexes keyed by their field name.
        field_cols = {c.value: c.column for c in header}

        # Get the field names.
        field_names = list(field_cols.keys())
        # Keep only the allowed field names.
        field_names.remove(refs_lbl)  # Remove the part reference field.
        cull_list(field_names, inc_field_names, exc_field_names)
        # Update the dictionary so it only has the allowed names.
        field_cols = {f: field_cols[f] for f in field_names}

        # Get the field values for each part reference.
        for row, ref in enumerate(refs, header_row + 1):
            if ref is None:
                continue  # Skip lines with no part reference.

            # Get the field values from the row of the current part reference.
            field_values = {}
            for field_name, col in list(field_cols.items()):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    field_values[field_name] = value
                else:
                    field_values[field_name] = ""

            # Explode the part reference into its individual references and
            # assign the field values to each part.
            for single_ref in explode(ref):
                part_fields[single_ref] = field_values

    except FindLabelError:
        logger.warn("No references column found.")
        raise FieldExtractionError

    if logger.isEnabledFor(DEBUG_DETAILED):
        print("Extracted Part Fields:")
        pprint(part_fields)

    return part_fields


def extract_part_fields_from_xlsx(
    filename, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from an XLSX spreadsheet."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from XLSX file {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    try:
        wb = pyxl.load_workbook(filename, data_only=True)
        return extract_part_fields_from_wb(wb, inc_field_names, exc_field_names)
    except FieldExtractionError:
        logger.warn("Field extraction failed on {}.".format(filename))
    return {}


def extract_part_fields_from_csv(
    filename, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from a CSV spreadsheet."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from CSV file {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    try:
        # Convert the CSV file into an XLSX workbook object and extract fields from that.
        wb, _ = csvfile_to_wb(filename)
        return extract_part_fields_from_wb(wb, inc_field_names, exc_field_names)
    except FieldExtractionError:
        logger.warn("Field extraction failed on {}.".format(filename))
    return {}


def extract_part_fields_from_sch(
    filename, inc_field_names=None, exc_field_names=None, recurse=False, depth=0
):
    """Return a dictionary of part fields extracted from a schematic."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from schematic file {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    part_fields_dict = {}  # Start with an empty part fields dictionary.

    sch = Schematic(filename)  # Read in the schematic.

    # Get all the part fields in the schematic and keep only the desired ones.
    # Remove the reference field (F0) from the list because that's used as as the dict key.
    field_names = sch.get_field_names()
    cull_list(field_names, None, ["reference"])
    cull_list(field_names, inc_field_names, exc_field_names)

    # Go through each component of the schematic, extracting its fields.
    for component in sch.components:

        # Get the fields and their values from the component.
        part_fields = {}
        for f in component.fields:
            value = unquote(f["ref"])
            name = unquote(f["name"])

            # Store the field and its value if the field name is in the list of
            # allowed fields.
            if name in field_names:
                part_fields[name] = value

        # Create a dictionary entry for each ref and assign the part fields to it.
        for ref in component.get_refs():
            if ref[0] == "#" or ref[-1] == "?":
                continue  # Skip pseudo-parts (e.g. power nets) and unallocated parts.

            # Some components (like resistor arrays) contain multiple units.
            # Add the fields for a unit to the part fields dict. By the end of
            # the loop, the part fields dict will have the union of the fields
            # of all of the units.
            part_fields.update(part_fields_dict.get(ref, {}))
            part_fields_dict[ref] = part_fields

    # If this schematic references other schematic sheets, then extract the part fields from those.
    if recurse:
        for sheet in sch.sheets:
            for field in sheet.fields:
                if field["id"] == "F1":
                    sheet_file = os.path.join(
                        os.path.dirname(filename), unquote(field["value"])
                    )
                    part_fields_dict.update(
                        extract_part_fields_from_sch(
                            sheet_file,
                            inc_field_names,
                            exc_field_names,
                            recurse,
                            depth + 1,
                        )
                    )
                    break

    # Print part fields for debugging if this is the top-level sheet of the schematic.
    if depth == 0:
        if logger.isEnabledFor(DEBUG_DETAILED):
            print("Extracted Part Fields:")
            pprint(part_fields_dict)

    return part_fields_dict


def extract_part_fields_from_sch_V6(
    filename, inc_field_names=None, exc_field_names=None, recurse=False, depth=0
):
    """Return a dictionary of part fields extracted from a schematic."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from schematic file {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    part_fields_dict = {}  # Start with an empty part fields dictionary.

    sch = Schematic_V6(filename)  # Read in the schematic.

    # Get all the part fields in the schematic and keep only the desired ones.
    # Remove the reference field (F0) from the list because that's used as as the dict key.
    field_names = sch.get_field_names()
    cull_list(field_names, None, ["reference"])
    cull_list(field_names, inc_field_names, exc_field_names)

    # Go through each component of the schematic, extracting its fields.
    for component in sch.components:

        # Get the fields and their values from the component.
        part_fields = {}
        for field in component.fields:
            # Store the field and its value if the field name is in the list of
            # allowed fields.
            name = field["name"]
            if name in field_names:
                part_fields[name] = field["value"]

        # Create a dictionary entry for the component ref and assign the part fields to it.
        ref = component.get_ref()
        if ref[0] == "#" or ref[-1] == "?":
            continue  # Skip pseudo-parts (e.g. power nets) and unallocated parts.

        # Some components (like resistor arrays) contain multiple units.
        # Add the fields for a unit to the part fields dict. By the end of
        # the loop, the part fields dict will have the union of the fields
        # of all of the units.
        part_fields.update(part_fields_dict.get(ref, {}))
        part_fields_dict[ref] = part_fields

    return part_fields_dict


def extract_part_fields_from_lib(
    filename, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from a library."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from part library {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    def get_field_names_lib(lib):
        """Return a list all the field names found in a library's components."""

        field_names = set(lib_field_id_to_name.values())
        field_names.add("prefix")
        for component in lib.components:
            for f in component.fields:
                try:
                    field_names.add(unquote(f["fieldname"]))
                except KeyError:
                    pass
        field_names.discard("")
        return list(field_names)

    part_fields_dict = {}  # Start with an empty part dictionary.

    lib = SchLib(filename)  # Read in all the parts in the library.

    # Get all the part fields in the schematic and keep only the desired ones.
    field_names = get_field_names_lib(lib)
    cull_list(field_names, inc_field_names, exc_field_names)

    # Go through each component in the library, extracting its fields.
    for component in lib.components:
        component_name = component.definition["name"]

        # Get the fields and their values from the component.
        part_fields = {}
        for id, f in enumerate(component.fields):
            if "reference" in list(f.keys()):
                name = "prefix"
                value = unquote(f["reference"])
            elif "name" in list(f.keys()):
                # Assign a name for the unnamed fields (F1, F2 & F3).
                # Use the already-assigned name for the higher fields (F4...).
                name = lib_field_id_to_name.get(str(id), unquote(f["fieldname"]))
                value = unquote(f["name"])
            else:
                logger.warn(
                    "Unknown type of field in part {}: {}.".format(component_name, f)
                )
                continue

            # Store the field and its value if the field name is in the list of
            # allowed fields.
            logger.log(
                DEBUG_OBSESSIVE,
                "Extracted library part: {} {} {}.".format(component_name, name, value),
            )
            if name in field_names:
                part_fields[name] = value

        # Create a dictionary entry for this library component.
        part_fields_dict[component_name] = part_fields

    if logger.isEnabledFor(DEBUG_DETAILED):
        print("Extracted Part Fields:")
        pprint(part_fields_dict)

    return part_fields_dict


def extract_part_fields_from_lib_V6(
    filename, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from a KiCad V6 library."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from part library {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    def get_field_names_lib(lib):
        """Return a list all the field names found in a library's components."""

        field_names = set(lib_field_id_to_name.values())
        field_names.add("prefix")
        for component in lib.components:
            for f in component.fields:
                try:
                    field_names.add(unquote(f["fieldname"]))
                except KeyError:
                    pass
        field_names.discard("")
        return list(field_names)

    part_fields_dict = {}  # Start with an empty part dictionary.

    lib = SchLib_V6(filename)  # Read in all the parts in the library.

    # Get all the part fields in the schematic and keep only the desired ones.
    field_names = lib.get_field_names()
    cull_list(field_names, inc_field_names, exc_field_names)

    # Go through each component in the library, extracting its fields.
    for component in lib.components:

        # Get the fields and their values from the component.
        part_fields = {}
        for field in component.fields:
            name = field["name"]
            value = field["value"]

            logger.log(
                DEBUG_OBSESSIVE,
                "Extracted library part: {} {} {}.".format(component.name, name, value),
            )

            # Store the field and its value if the field name is in the list of
            # allowed fields.
            if name in field_names:
                part_fields[name] = value

        # Create a dictionary entry for this library component.
        part_fields_dict[component.name] = part_fields

    if logger.isEnabledFor(DEBUG_DETAILED):
        print("Extracted Part Fields:")
        pprint(part_fields_dict)

    return part_fields_dict


def extract_part_fields_from_dcm(
    filename, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from a part description file."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields {}, -{} from part description file {}.".format(
            inc_field_names, exc_field_names, filename
        ),
    )

    part_fields_dict = {}  # Start with an empty part dictionary.

    try:
        dcm = Dcm(filename)
    except IOError:
        return part_fields_dict  # Return empty part fields dict if no DCM file found.

    # Start with DCM field names and keep the desired ones.
    field_names = deepcopy(dcm_field_names)
    cull_list(field_names, inc_field_names, exc_field_names)

    # Go through each component, extracting its fields.
    for component in dcm.components:
        component_name = component.name

        # Get the fields and their values from the component.
        part_fields = {}
        for name in field_names:
            value = getattr(component, name, None)
            if value is not None:
                logger.log(
                    DEBUG_OBSESSIVE,
                    "Extracted part description: {} {} {}.".format(
                        component_name, name, value
                    ),
                )
                part_fields[name] = value

        # Create a dictionary entry for this library component.
        part_fields_dict[component_name] = part_fields

    if logger.isEnabledFor(DEBUG_DETAILED):
        print("Extracted Part Fields:")
        pprint(part_fields_dict)

    return part_fields_dict


def combine_part_field_dicts(from_dict, to_dict, do_union=True):
    """Combine two part field dictionaries."""

    if to_dict is None or len(to_dict) == 0:
        comb_dict = deepcopy(from_dict)
    else:
        comb_dict = deepcopy(to_dict)
        to_dict_part_refs = list(to_dict.keys())

        # Go through the parts in the FROM dictionary...
        for from_ref, from_fields in list(from_dict.items()):

            # If the TO dictionary has the same part...
            if from_ref in to_dict_part_refs:
                # Then insert the fields in the FROM part into the TO part.
                comb_dict[from_ref].update(from_fields)

            # If the FROM part isn't in the TO dictionary, but a union operation is active...
            elif do_union:
                # Then add the FROM part into the TO dictionary.
                comb_dict[from_ref] = from_fields

    return comb_dict


def extract_part_fields(
    filenames, inc_field_names=None, exc_field_names=None, recurse=False
):
    """Return a dictionary of part fields extracted from a spreadsheet, part library, DCM, or schematic."""

    logger.log(
        DEBUG_OVERVIEW,
        "Extracting fields +{}, -{} from files {}.".format(
            inc_field_names, exc_field_names, filenames
        ),
    )

    # Table of extraction functions for each file type.
    extraction_functions = {
        ".xlsx": extract_part_fields_from_xlsx,
        ".tsv": extract_part_fields_from_csv,
        ".csv": extract_part_fields_from_csv,
        ".sch": extract_part_fields_from_sch,
        ".kicad_sch": extract_part_fields_from_sch_V6,
        ".lib": extract_part_fields_from_lib,
        ".kicad_sym": extract_part_fields_from_lib_V6,
        ".dcm": extract_part_fields_from_dcm,
    }

    part_fields_dict = {}  # Start with empty part field dictionary.

    # If extracting from only a single file, make a one-entry list.
    if type(filenames) == str:
        filenames = [filenames]

    # Extract the fields from the parts in each file.
    for f in filenames:
        try:
            logger.log(DEBUG_DETAILED, "Extracting fields from {}.".format(f))

            # Set the extraction function based on the file extension.
            f_extension = os.path.splitext(f)[1].lower()
            extraction_function = extraction_functions[f_extension]

        except KeyError:
            logger.warn("Unknown file type for field extraction: {}.".format(f))

        else:
            # Call the extraction function.
            try:
                f_part_fields_dict = extraction_function(
                    f, inc_field_names, exc_field_names, recurse
                )

            except IOError:
                logger.warn("File not found: {}.".format(f))

            else:
                # Add the extracted fields to the total part dictionary.
                part_fields_dict = combine_part_field_dicts(
                    f_part_fields_dict, part_fields_dict
                )

    if logger.isEnabledFor(DEBUG_DETAILED):
        print("Total Extracted Part Fields:")
        pprint(part_fields_dict)

    if part_fields_dict is None or len(part_fields_dict) == 0:
        logger.warn(
            "No part fields were extracted from these files: {}\n  * Did you provide a list of files?\n  * Do these files exist?\n  * Did you annotate your components in the schematic?".format(
                ", ".join(filenames)
            )
        )
        return

    return part_fields_dict


def insert_part_fields_into_wb(part_fields_dict, wb, recurse=False):
    """Insert the fields in the extracted part dictionary into an XLSX workbook."""

    id_label = "Refs"

    # Get all the unique field labels used in the dictionary of part fields.
    field_labels = set([])
    for fields_and_values in part_fields_dict.values():
        for field_label in fields_and_values:
            field_labels.add(field_label)
    field_labels = sorted(field_labels)
    field_labels.insert(0, id_label)

    if wb is None:
        wb = pyxl.Workbook()  # No workbook given, so create one.

    wb = ungroup_wb(wb)  # Ungroup any grouped references.

    ws = wb.active  # Get the active sheet from the workbook.

    def set_cell_format(cell):
        """Set cell to TEXT format if it contains a string."""
        if cell.data_type == "s":
            cell.number_format = "@"

    if ws.min_column == ws.max_column:
        # If the given worksheet is empty, then create one using the part field labels.
        # Create header row with a column for each part field.
        for c, lbl in enumerate(field_labels, 1):
            cell = ws.cell(row=1, column=c)
            cell.value = lbl
            set_cell_format(cell)

        # Enter the part references into the part reference column.
        for row, ref in enumerate(sorted(part_fields_dict.keys()), 2):
            cell = ws.cell(row=row, column=1)
            cell.value = ref
            set_cell_format(cell)

    # Get the header row from the worksheet.
    header_row, headers = find_header(ws)
    header_labels = [cell.value for cell in headers]

    # Get column for each header field.
    header_columns = {h.value: h.column for h in headers}

    # Next open column. Combine with [0] in case there are no headers.
    next_header_column = max(list(header_columns.values()) + [0]) + 1

    # Find the column with the part references.
    ref_col, _ = find_header_column(headers, id_label)

    # Add all the missing part references from the field dictionary to the worksheet.
    # That will ensure a worksheet that only had a subset of the dictionary
    # part fields will get the missing ones.
    refs = set([])
    for row in range(header_row + 1, ws.max_row + 1):
        for ref in explode(ws.cell(row=row, column=ref_col).value):
            refs.add(ref)
    row = ws.max_row + 1
    for ref in sorted(part_fields_dict.keys()):
        if ref not in refs:
            cell = ws.cell(row=row, column=ref_col)
            cell.value = ref
            set_cell_format(cell)
            refs.add(ref)
            row += 1

    # Go through each row, see if any reference is in the part dictionary, and
    # insert/overwrite fields from the dictionary.
    for row in range(header_row + 1, ws.max_row + 1):
        for ref in explode(ws.cell(row=row, column=ref_col).value):
            try:
                fields = part_fields_dict[ref]
                for field, value in fields.items():
                    # Skip None fields.
                    if value is None:
                        continue

                    try:
                        # Match the field name to one of the headers and overwrite the
                        # cell value with the dictionary value.
                        header = lc_get_close_matches(field, header_labels, 1, 0.3)[0]
                        cell = ws.cell(row=row, column=header_columns[header])
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Updating {} field {} from {} to {}".format(
                                ref, field, cell.value, value
                            ),
                        )
                        cell.value = value
                        set_cell_format(cell)
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Type of {} field {} containing {} is {}".format(
                                ref, field, cell.value, cell.data_type
                            ),
                        )

                    except IndexError:
                        # The dictionary field didn't match any sheet header closely enough,
                        # so add a new column with the field name as the header label.
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Adding {} field {} with value {}".format(
                                ref, field, value
                            ),
                        )
                        cell = ws.cell(row=row, column=next_header_column)
                        cell.value = value
                        set_cell_format(cell)
                        new_header_cell = ws.cell(
                            row=header_row, column=next_header_column
                        )
                        new_header_cell.value = field
                        headers.append(new_header_cell)
                        header_labels.append(field)
                        header_columns[field] = next_header_column
                        next_header_column += 1

            except KeyError:
                pass  # The part reference doesn't exist in the dictionary.

    return wb


def insert_part_fields_into_xlsx(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into an XLSX spreadsheet."""

    logger.log(
        DEBUG_OVERVIEW, "Inserting extracted fields into XLSX file {}.".format(filename)
    )

    if backup:
        create_backup(filename)

    # Either insert fields into an existing workbook, or use an empty one.
    try:
        wb = pyxl.load_workbook(filename, data_only=True)
    except IOError:
        wb = None

    wb = insert_part_fields_into_wb(part_fields_dict, wb)

    if group_components:
        wb = group_wb(wb, no_range)

    wb.save(filename)


def insert_part_fields_into_csv(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a CSV spreadsheet."""

    logger.log(
        DEBUG_OVERVIEW, "Inserting extracted fields into CSV file {}.".format(filename)
    )

    if backup:
        create_backup(filename)

    # Either insert fields into an existing workbook, or use an empty one.
    try:
        wb, dialect = csvfile_to_wb(filename)
    except IOError:
        wb = None
        if os.path.splitext(filename)[-1] == ".tsv":
            dialect = "excel-tab"
        else:
            dialect = "excel"

    wb = insert_part_fields_into_wb(part_fields_dict, wb)

    if group_components:
        wb = group_wb(wb, no_range)

    wb_to_csvfile(wb, filename, dialect)


def insert_part_fields_into_sch(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a schematic."""

    logger.log(
        DEBUG_OVERVIEW,
        "Inserting extracted fields into schematic file {}.".format(filename),
    )

    def reorder_sch_fields(fields):
        """Return the part fields with the named fields ordered alphabetically."""
        # Sort the named fields that come after the first four, unnamed fields.
        sort_key = operator.itemgetter("name")
        if USING_PYTHON2:
            sort_key_func = lambda s: unicode(sort_key(s))
        else:
            sort_key_func = sort_key
        named_fields = sorted(fields[4:], key=sort_key_func)
        # Renumber the ids of the sorted fields.
        for id, field in enumerate(named_fields, 4):
            field["id"] = str(id)
        # Return the first four fields plus the remaining sorted fields.
        return fields[:4] + named_fields

    # Get an existing schematic or abort. (There's no way we can create
    # a viable schematic file just from part field values.)
    try:
        sch = Schematic(filename)
    except IOError:
        logger.warn("Schematic file {} not found.".format(filename))
        return

    # Go through all the schematic components, replacing field values and
    # adding new fields found in the part fields dictionary.
    for component in sch.components:

        prev_part_fields = None

        # For each reference for this component, search in the dictionary
        # for new or updated fields for this part.
        refs = component.get_refs()
        for ref in refs:

            # Get the part fields for the given part reference (or an empty list).
            part_fields = part_fields_dict.get(ref, {})

            # Warn if the current part fields for this component don't match the
            # previous part fields (which may happen with hierarchical schematics).
            if prev_part_fields is not None and part_fields != prev_part_fields:
                logger.warn(
                    "The inserted part lists for hierarchically-instantiated components {} have different values.".format(
                        refs
                    )
                )
            # Store the part fields for later comparison.
            prev_part_fields = deepcopy(part_fields)

            # Insert the fields from the part dictionary into the component fields.
            for field_name, field_value in part_fields.items():

                # Create a dict to hold the field visibility attribute.
                try:
                    field_attributes = dict()
                    INVIS_CODE = "0001"
                    VISIBLE_CODE = "0000"
                    if field_name.startswith(INVIS_PREFIX):
                        field_attributes["attributes"] = INVIS_CODE
                        field_name = field_name[len(INVIS_PREFIX) :]
                    elif field_name.startswith(VISIBLE_PREFIX):
                        field_attributes["attributes"] = VISIBLE_CODE
                        field_name = field_name[len(VISIBLE_PREFIX) :]
                    if field_value.startswith(INVIS_PREFIX):
                        field_attributes["attributes"] = INVIS_CODE
                        field_value = field_value[len(INVIS_PREFIX) :]
                    elif field_value.startswith(VISIBLE_PREFIX):
                        field_attributes["attributes"] = VISIBLE_CODE
                        field_value = field_value[len(VISIBLE_PREFIX) :]
                except AttributeError:
                    # If we get here, it's probably because field_value is not a
                    # string so the startswith() method wasn't found. Because it's
                    # not a string, there's no way for it to have a prefix string
                    # so we can just ignore the exception because the action never
                    # would have happened anyway.
                    pass

                # Also store a position for a new field based on the REF position.
                posx = component.fields[0]["posx"]
                posy = str(
                    int(component.fields[0]["posy"]) + 100
                )  # Place it below REF.
                field_position = {"posx": posx, "posy": posy}

                # Get the field id associated with this field name (if there is one).
                field_id = lib_field_name_to_id.get(field_name, None)

                # Search for an existing field with a matching name in the component.
                for f in component.fields:

                    if unquote(f["name"]).lower() == field_name.lower():
                        # Update existing named field in component.
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Updating {} field {} from {} to {}".format(
                                ref, f["id"], f["ref"], quote(field_value)
                            ),
                        )
                        f["ref"] = quote(field_value)
                        # Set field attributes but don't change its position.
                        if "attributes" in field_attributes:
                            f["attributes"] = field_attributes["attributes"]
                        break

                    elif f["id"] == field_id:
                        # Update one of the default, unnamed fields in component.
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Updating {} field {} from {} to {}".format(
                                ref, f["id"], f["ref"], quote(field_value)
                            ),
                        )
                        f["ref"] = quote(field_value)
                        # Set field attributes but don't change its position.
                        if "attributes" in field_attributes:
                            f["attributes"] = field_attributes["attributes"]
                        break

                # No existing field to update, so add a new field.
                else:
                    if field_value not in (None, ""):
                        # Add new named field and value to component.
                        new_field = {
                            "ref": quote(field_value),
                            "name": quote(field_name),
                        }
                        new_field.update(field_attributes)  # Set field's attributes.
                        new_field.update(field_position)  # Set new field's position.
                        component.add_field(new_field)
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Adding {} field {} with value {}".format(
                                ref, component.fields[-1]["id"], quote(field_value)
                            ),
                        )

                # Keep only default fields and named fields with non-empty values.
                default_field_ids = sch_field_id_to_name.keys()
                component.fields = [
                    f
                    for f in component.fields
                    if f["id"] in default_field_ids
                    or unquote(f.get("ref", None)) not in (None, "")
                ]

                # Canonically order the fields to make schematic comparisons
                # easier during acceptance testing.
                component.fields = reorder_sch_fields(component.fields)

    # Save the updated schematic.
    if backup:
        create_backup(filename)
    sch.save(filename)

    # If this schematic references other schematic sheets, then insert the part fields into those, too.
    if recurse:
        for sheet in sch.sheets:
            # If filename includes a path, save this path to prepend below
            if filename.count("/") > 0:
                prepend_dir = filename.rsplit("/", 1)[0] + "/"
            else:
                prepend_dir = "./"
            for field in sheet.fields:
                if field["id"] == "F1":
                    # Prepend path for sheets which are nested more than once
                    sheet_file = prepend_dir + unquote(field["value"])
                    insert_part_fields_into_sch(
                        part_fields_dict, sheet_file, recurse, group_components, backup, no_range
                    )
                    break


def insert_part_fields_into_sch_V6(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a schematic."""

    logger.log(
        DEBUG_OVERVIEW,
        "Inserting extracted fields into schematic file {}.".format(filename),
    )

    # Get an existing schematic or abort. (There's no way we can create
    # a viable schematic file just from part field values.)
    try:
        sch = Schematic_V6(filename)
    except IOError:
        logger.warn("Schematic file {} not found.".format(filename))
        return

    # Go through all the schematic components, replacing field values and
    # adding new fields found in the part fields dictionary.
    for component in sch.components:

        prev_part_fields = None

        # For each reference for this component, search in the dictionary
        # for new or updated fields for this part.
        ref = component.get_ref()

        # Get the part fields for the given part reference (or an empty list).
        part_fields = part_fields_dict.get(ref, {})

        # Warn if the current part fields for this component don't match the
        # previous part fields (which may happen with hierarchical schematics).
        if prev_part_fields is not None and part_fields != prev_part_fields:
            logger.warn(
                "The inserted part lists for hierarchically-instantiated components {} have different values.".format(
                    refs
                )
            )
        # Store the part fields for later comparison.
        prev_part_fields = deepcopy(part_fields)

        # Insert the fields from the part dictionary into the component fields.
        for field_name, field_value in part_fields.items():

            # Get [V] (visible) or [I] (invisible) flag prepended to entire field name
            # or individual field value. If no flag, visibility is set to None.
            field_vis, field_name = re.match(r"(\[([VI])\])?(.*)", field_name).group(
                2, 3
            )
            value_vis, field_value = re.match(r"(\[([VI])\])?(.*)", field_value).group(
                2, 3
            )
            if value_vis == "V":
                total_vis = True
            elif value_vis == "I":
                total_vis = False
            else:
                if field_vis == "V":
                    total_vis = True
                elif field_vis == "I":
                    total_vis = False
                else:
                    total_vis = None

            # Search for an existing field with a matching name in the component.
            for f in component.fields:

                if f["name"].lower() == field_name.lower():
                    # Update existing named field in component.
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Updating {} field {} from {} to {} with visibility {}".format(
                            ref, f["name"], f["value"], quote(field_value), total_vis
                        ),
                    )
                    f["value"] = field_value
                    component.set_field_value(f["name"], field_value)
                    component.set_field_visibility(f["name"], total_vis)
                    break  # To keep following else from creating a new field.

            # No existing field to update, so add a new field.
            else:
                if field_value not in (None, ""):
                    # Add new named field and value to component.
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Adding {} field {} with value {} with visibility {}".format(
                            ref, field_name, quote(field_value), total_vis
                        ),
                    )
                    component.copy_field("Reference", field_name)
                    component.set_field_value(field_name, field_value)
                    pos = component.get_field_pos(field_name)
                    field = component.get_field(field_name)
                    pos[1] += 2.54 * field["id"]
                    component.set_field_pos(field_name, pos)
                    if total_vis is None or total_vis == False:
                        component.set_field_visibility(field_name, False)

        # Remove non-default fields with empty values.
        for field in component.fields:
            name = field["name"]
            if name.lower() in ("reference", "value", "footprint", "datasheet"):
                # Skip default fields so they aren't removed.
                continue
            if field["value"] in (None, ""):
                # Remove empty field.
                component.del_field(name)

    # Save the updated schematic and sub-schematics (if recursing).
    sch.save(recurse, backup, filename)


def insert_part_fields_into_lib(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a library."""

    logger.log(
        DEBUG_OVERVIEW,
        "Inserting extracted fields into library file {}.".format(filename),
    )

    if backup:
        create_backup(filename)

    # Get an existing library or abort. (There's no way we can create
    # a viable library file just from part field values.)
    try:
        lib = SchLib(filename)
    except IOError:
        logger.warn("Library file {} not found.".format(filename))
        return

    # Go through all the library components, replacing field values and
    # adding new fields from the part fields dictionary.
    for component in lib.components:
        component_name = component.definition["name"]

        # Get fields for the part with the same name as this component (or an empty list).
        part_fields = part_fields_dict.get(component_name, {})

        # Insert the fields from the part dictionary into the component fields.
        for field_name, field_value in part_fields.items():

            # Get the field id associated with this field name (if there is one).
            field_id = lib_field_name_to_id.get(field_name, None)

            # Search for an existing field with a matching name in the component.
            for id, f in enumerate(component.fields):

                if unquote(f.get("fieldname", "")).lower() == field_name.lower():
                    # Update existing named field in component.
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Updating {} field {} from {} to {}".format(
                            component_name, field_name, f["name"], quote(field_value)
                        ),
                    )
                    f["name"] = quote(field_value)
                    break

                elif str(id) == field_id:
                    if id == 0:
                        # Update the F0 field of the component.
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Updating {} field {} from {} to {}".format(
                                component_name,
                                field_id,
                                f["reference"],
                                quote(field_value),
                            ),
                        )
                        f["reference"] = quote(field_value)
                    else:
                        # Update one of the F1, F2, or F3 fields in the component.
                        logger.log(
                            DEBUG_OBSESSIVE,
                            "Updating {} field {} from {} to {}".format(
                                component_name, field_id, f["name"], quote(field_value)
                            ),
                        )
                        f["name"] = quote(field_value)
                    break

            # No existing field to update, so add a new field.
            else:
                if field_value not in (None, ""):
                    # Copy an existing field from the component and then
                    # update its name and value to create a new field.
                    new_field = deepcopy(component.fields[-1])
                    new_field["fieldname"] = quote(field_name)
                    new_field["name"] = quote(field_value)
                    component.fields.append(new_field)
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Adding {} field {} with value {}".format(
                            component_name, field_name, quote(field_value)
                        ),
                    )

        # Remove any named fields with empty values.
        component.fields = [
            f
            for f in component.fields
            if unquote(f.get("fieldname", None)) in (None, "", "~")
            or unquote(f.get("name", None)) not in (None, "")
        ]

    # Save the updated library.
    lib.save(filename)


def insert_part_fields_into_lib_V6(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a KiCad V6 library."""

    logger.log(
        DEBUG_OVERVIEW,
        "Inserting extracted fields into library file {}.".format(filename),
    )

    # Get an existing library or abort. (There's no way we can create
    # a viable library file just from part field values.)
    try:
        lib = SchLib_V6(filename)
    except IOError:
        logger.warn("Library file {} not found.".format(filename))
        return

    # Go through all the library components, replacing field values and
    # adding new fields from the part fields dictionary.
    for component in lib.components:

        # Get fields for the part with the same name as this component (or an empty list).
        part_fields = part_fields_dict.get(component.name, {})

        # Insert the fields from the part dictionary into the component fields.
        for field_name, field_value in part_fields.items():

            # Get [V] (visible) or [I] (invisible) flag prepended to entire field name
            # or individual field value. If no flag, visibility is set to None.
            field_vis, field_name = re.match(r"(\[([VI])\])?(.*)", field_name).group(
                2, 3
            )
            value_vis, field_value = re.match(r"(\[([VI])\])?(.*)", field_value).group(
                2, 3
            )
            if value_vis == "V":
                total_vis = True
            elif value_vis == "I":
                total_vis = False
            else:
                if field_vis == "V":
                    total_vis = True
                elif field_vis == "I":
                    total_vis = False
                else:
                    total_vis = None

            # Search for an existing field with a matching name in the component.
            for f in component.fields:

                if f["name"].lower() == field_name.lower():
                    # Update existing named field in component.
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Updating {} field {} from {} to {}".format(
                            component.name, f["name"], f["value"], quote(field_value)
                        ),
                    )
                    f["value"] = field_value
                    component.set_field_value(f["name"], field_value)
                    component.set_field_visibility(f["name"], total_vis)
                    break  # To keep following else from creating a new field.

            # No existing field to update, so add a new field.
            else:
                if field_value not in (None, ""):
                    # Add new named field and value to component.
                    logger.log(
                        DEBUG_OBSESSIVE,
                        "Adding {} field {} with value {} with visibility {}".format(
                            component.name, f["name"], f["value"], total_vis
                        ),
                    )
                    component.copy_field("Reference", field_name)
                    component.set_field_value(field_name, field_value)
                    pos = component.get_field_pos(field_name)
                    field = component.get_field(field_name)
                    pos[1] += 2.54 * field["id"]
                    component.set_field_pos(field_name, pos)
                    if total_vis is None or total_vis == False:
                        component.set_field_visibility(field_name, False)

        # Remove non-default fields with empty values.
        for field in component.fields:
            name = field["name"]
            if name.lower() in (
                "reference",
                "value",
                "footprint",
                "datasheet",
                "ki_description",
                "ki_fp_filters",
                "ki_keywords",
                "ki_locked",
            ):
                # Skip default fields so they aren't removed.
                continue
            if field["value"] in (None, ""):
                # Remove empty field.
                component.del_field(name)

    # Save the updated library.
    lib.save(backup, filename)


def insert_part_fields_into_dcm(
    part_fields_dict, filename, recurse, group_components, backup, no_range
):
    """Insert the fields in the extracted part dictionary into a DCM file."""

    logger.log(
        DEBUG_OVERVIEW, "Inserting extracted fields into DCM file {}.".format(filename)
    )

    if backup:
        create_backup(filename)

    # Get the part fields from the DCM file.
    dcm_part_fields_dict = extract_part_fields_from_dcm(filename)

    # Add the part fields from the part field dictionary.
    dcm_part_fields_dict = combine_part_field_dicts(
        part_fields_dict, dcm_part_fields_dict
    )

    # Create a new Dcm object from the combined part fields.
    dcm = Dcm()
    for part_name, fields in list(dcm_part_fields_dict.items()):
        cmp = Component()
        cmp.name = part_name
        for k, v in list(fields.items()):
            if k in dcm_field_names:
                setattr(cmp, k, v)
        dcm.components.append(cmp)

    # Overwrite the current DCM file with the new part fields.
    dcm.save(filename)


def insert_part_fields(part_fields_dict, filenames, recurse, group_components, backup, no_range):
    """Insert part fields from a dictionary into a spreadsheet, part library, or schematic."""

    # No files backed-up yet, so clear list of file names.
    global backedup_files
    backedup_files = []

    logger.log(
        DEBUG_OVERVIEW, "Inserting extracted fields into files {}.".format(filenames)
    )

    # Table of insertion functions for each file type.
    insertion_functions = {
        ".xlsx": insert_part_fields_into_xlsx,
        ".tsv": insert_part_fields_into_csv,
        ".csv": insert_part_fields_into_csv,
        ".sch": insert_part_fields_into_sch,
        ".kicad_sch": insert_part_fields_into_sch_V6,
        ".lib": insert_part_fields_into_lib,
        ".kicad_sym": insert_part_fields_into_lib_V6,
        ".dcm": insert_part_fields_into_dcm,
    }

    if part_fields_dict is None or len(part_fields_dict) == 0:
        logger.warn("There are no part field values to insert!")
        return

    # If inserting into a single file, make a one-entry list.
    if type(filenames) == str:
        filenames = [filenames]

    # Insert the part fields into each file.
    for f in filenames:
        try:
            logger.log(DEBUG_DETAILED, "Inserting fields into {}.".format(f))

            # Set the insertion function based on the file extension.
            f_extension = os.path.splitext(f)[1].lower()
            insertion_function = insertion_functions[f_extension]

        except KeyError:
            logger.warn("Unknown file type for field insertion: {}".format(f))

        else:
            try:
                insertion_function(
                    part_fields_dict, f, recurse, group_components, backup, no_range
                )

            except IOError:
                logger.warn("Unable to write to file: {}.".format(f))


def clean_part_fields(part_fields_dict):
    """Clean field values (i.e., remove or replace any newlines with spaces.)"""
    for part, fields in part_fields_dict.items():
        for k, v in fields.items():
            v = re.sub("[\n\r]+$", "", str(v))  # Remove newlines at end of field.
            fields[k] = re.sub("[\n\r]+", " ", v)  # Replace newlines within field.


def kifield(
    extract_filenames,
    insert_filenames,
    inc_field_names=None,
    exc_field_names=None,
    recurse=False,
    group_components=False,
    backup=True,
    no_range=False
):
    """Extract fields from a set of files and insert them into another set of files."""

    # Extract a dictionary of part field values from a set of files.
    part_fields_dict = extract_part_fields(
        extract_filenames, inc_field_names, exc_field_names, recurse
    )

    clean_part_fields(part_fields_dict)

    # Insert entries from the dictionary into these files.
    insert_part_fields(
        part_fields_dict, insert_filenames, recurse, group_components, backup, no_range
    )
